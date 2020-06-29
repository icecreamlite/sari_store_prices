#!/home/b/projects/sari_store_prices/venv/bin/python3

#link for openpyxl tutorial
#https://code.tutsplus.com/tutorials/how-to-work-with-excel-documents-using-python--cms-25698

#Link for pandas excel tutorial
#https://pythonbasics.org/read-excel/
#https://pythonbasics.org/write-excel/


from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, messagebox
from tkinter import font as tkFont
import re, pandas
from sys import path

sspDir = '/home/b/projects/sari_store_prices/'

path.append(sspDir + 'modules')

from tagloc_generate import generateTagLocFile

generateTagLocFile(sspDir)

list_excel_rows = [] #list for excel row number of displayed items
selected_line = 0 #initialize line number of selected item
selected_row = [None, None, None, None]
delBool = False
histFirstOpen = True

excel_file = load_workbook(sspDir + 'ssp.xlsx')
sheet1 = excel_file.get_sheet_by_name('Sheet1')
num_items = sheet1.max_row #get number of rows

col_title = '-' * 59 + '\n' + '||' + 'ITEM'.center(46) + '||' + 'PRICE'.center(7) + '||'\
                + '\n' + '-' * 59 + '\n'


def reloadExcel():
    global excel_file
    excel_file = load_workbook(sspDir + 'ssp.xlsx')
    global sheet1
    sheet1 = excel_file.get_sheet_by_name('Sheet1')
    global num_items
    num_items = sheet1.max_row

def sortExcel():
    df = pandas.read_excel(sspDir + 'ssp.xlsx')
    df = df.sort_values('Item')
    df.to_excel(sspDir + 'ssp.xlsx', index=False)

def updateText():
    #check if filters are applied
    filter_tag = False
    filter_loc = False

    selected_tag = tagsVar.get()
    if selected_tag != 'Tags':
        filter_tag = True

    selected_loc = locVar.get()
    if selected_loc != 'Locations':
        filter_loc = True

    #used to show all match to filters if entry is empty
    typed = ent.get()
    if filter_tag or filter_loc:
        if typed == '':
            typed = '.'

    #find all match to filters and entry
    list_text = ''
    global list_excel_rows
    list_excel_rows = []
    if typed != '':
        regex = re.compile(typed, re.I)
        for i in range(1, num_items):
            regex_matched = False
            if regex.search(str(sheet1['A' + str(i+1)].value)) != None:
                regex_matched = True
            if filter_tag:
                if str(sheet1['C' + str(i+1)].value) != selected_tag:
                    regex_matched = False
            if filter_loc:
                if str(sheet1['D' + str(i+1)].value) != selected_loc:
                    regex_matched = False
            if regex_matched:
                list_excel_rows.append(str(i+1))
                list_text += '||' + str(sheet1['A' + str(i+1)].value).center(46) + '||' + str(sheet1['B' + str(i+1)].value).center(7) + '||\n'
        try:
            list_excel_rows[0]
            list_text += '-' * 59
        except IndexError:
            pass
    #update tkinter price_list text box based on matches
    price_list['state'] = 'normal'
    price_list.delete(1.0, 'end')
    price_list.insert(1.0, col_title + list_text)
    price_list['state'] = 'disabled'


#edit or add item
def submitItem():
    trig = edit_add.get()
    updateBool = True
    same_name = False
    iE_bg = 'white'
    pE_bg = 'white'
    Cbtheme = 'white'

    item = itemEntry.get()
    price = priceEntry.get()
    tag = tagCb.get()
    location = locationCb.get()

    if trig == 'edit':
        row = list_excel_rows[selected_line - 4]
        tracker_msg = f'Edited row {list_excel_rows[selected_line - 4]}: {selected_row[0]}|{selected_row[1]}|{selected_row[2]}|{selected_row[3]} => {item}|{price}|{tag}|{location}\n'

        #if same data in entry, exit and do nothing
        if selected_row[0] == item:
            same_name = True
            if selected_row[1] == str(price) and selected_row[2] == tag and selected_row[3] == location:
                editWindowWithdraw()

    elif trig == 'add':
        same_name = False
        global num_items
        row = str(num_items + 1)
        tracker_msg = f'Inserted: {item}|{price}|{tag}|{location}\n'

    #check validity and update excel file if editWindow is still open
    if editWindow.state() == 'normal':
        try:
            price = int(price)
            assert price > 0
        except:
            try:
                price = float(price)
                assert price > 0
            except:
                updateBool = False
                pE_bg = 'red'

        #checks if item entry is empty (invalid input)
        if item == '':
            updateBool = False
            iE_bg = 'red'
        if tag == '' or location == '':
            updateBool = False
            Cbtheme = 'red'

        #update options based on error occurence in input fields
        itemEntry['bg'] = iE_bg
        priceEntry['bg'] = pE_bg
        style.theme_use(Cbtheme)

        #if both entry fields are valid, update the file
        #save changes to tracker text file
        #close editWindow
        if updateBool:
            sheet1['A' + row] = item
            sheet1['B' + row] = price
            sheet1['C' + row] = tag
            sheet1['D' + row] = location
            excel_file.save(sspDir + 'ssp.xlsx')

            #track changes
            with open(sspDir + 'text_files/change_tracker.txt', 'a') as ct:
                ct.write(tracker_msg)

            editWindowWithdraw()

            if not same_name:
                sortExcel()

            #generate tag and location text files only if they're updated
            if selected_row[2] != tag or selected_row[3] != location:
                generateTagLocFile(sspDir)


            updateAll(tracker_msg)


def editWindowWithdraw():
    ent.focus_set()
    editWindow.withdraw()
    tagsOM['state'] = 'normal'
    locationsOM['state'] = 'normal'
    addBut['state'] = 'normal'
    ent['state'] = 'normal'
    itemEntry['bg'] = 'white'
    priceEntry['bg'] = 'white'
    style.theme_use('white')


def fetchTags():
    tags = []
    for tag in open(sspDir + 'text_files/tags.txt', 'r'):
        tag = tag[:-1]
        tags.append(tag)
    return tags

def fetchLocs():
    locations = []
    for location in open(sspDir + 'text_files/locations.txt', 'r'):
        location = location[:-1]
        locations.append(location)
    return locations


def updateAll(tracker_msg): #Update GUI live
    displayHistLine(tracker_msg)
    reloadExcel()
    global tags
    tags = fetchTags()
    global locations
    locations = fetchLocs()
    updateFilters()
    updateText()


def showEditWindow(trig):
    tagsOM['state'] = 'disabled'
    locationsOM['state'] = 'disabled'
    edit_add.set(trig)
    tagCb.delete(0,'end')
    locationCb.delete(0, 'end')
    addBut['state'] = 'disabled'
    ent['state'] = 'disabled'    
    tagCb['value'] = tags
    locationCb['value'] = locations
    itemEntry.delete(0, 'end')
    priceEntry.delete(0, 'end')

    if trig == 'edit':
        editWindow.title('Edit Item')
        tagCurInd = tags.index(selected_row[2])
        locCurInd = locations.index(selected_row[3])

        #-- Set editWindow Fields
        #setEditWindowFields()
        tagCb.current(tagCurInd)
        locationCb.current(locCurInd)                
        itemEntry.insert(END, selected_row[0])
        priceEntry.insert(END, selected_row[1])
        #--
    elif trig == 'add': editWindow.title('Add Item')

    editWindow.attributes('-topmost', 'true')
    editWindow.deiconify()    


def mapKey(event):
    global delBool
    #do if edit window or delete window is not active
    if editWindow.state() == 'withdrawn' and delBool == False:
        current_focus = str(root.focus_get())

        if current_focus == '.!frame.!entry':
            if event.keysym == 'Escape': #clear entry and reset filters
                tagsVar.set('Tags')
                locVar.set('Locations')
                ent.delete(0, 'end')

            updateText()

        elif current_focus == '.!frame2.!text':

            if event.keysym == 'Escape':
                ent.focus_set()
                updateText()

            if event.keysym == 'Return':

                showEditWindow('edit')
                
    
    if delBool:
        delBool = False


def selectItem(event):
    if editWindow.state() == 'withdrawn':
        line_num = int(float(event.widget.index(CURRENT)))

        #if selected, store values and highlight
        if line_num > 3 and line_num < len(list_excel_rows) + 4:
            global selected_line
            selected_line = line_num
            global selected_row
            selected_row[0] = str(sheet1['A' + list_excel_rows[selected_line - 4]].value) #Item name
            selected_row[1] = str(sheet1['B' + list_excel_rows[selected_line - 4]].value) #Price
            selected_row[2] = str(sheet1['C' + list_excel_rows[selected_line - 4]].value) #Tag
            selected_row[3] = str(sheet1['D' + list_excel_rows[selected_line - 4]].value) #Location
            event.widget.focus_set()
            updateText()
            event.widget.tag_add('selected', float(line_num), line_num + 0.59)
            event.widget.tag_config('selected', background='DodgerBlue2')


def delItem(event):
    global delBool
    delBool = True
    ent['state'] = 'disabled'
    ans = messagebox.askyesno('Delete', f"Are you sure you want to delete {selected_row[0]}?")
    if ans == True:
        ent.focus_set()

        #track deletion
        tracker_msg = f'Deleted row {list_excel_rows[selected_line - 4]}: {selected_row[0]}|{selected_row[1]}|{selected_row[2]}|{selected_row[3]}\n'
        with open(sspDir + 'text_files/change_tracker.txt', 'a') as ct:
            ct.write(tracker_msg)

        #delete row and save
        sheet1.delete_rows(int(list_excel_rows[selected_line - 4]), 1)
        excel_file.save(sspDir + 'ssp.xlsx')

        generateTagLocFile(sspDir)
        
        updateAll(tracker_msg)

    ent.focus_set()
    ent['state'] = 'normal'


def setFilter(filter, value):
    if filter == 'tag':
        tagsVar.set(value)
    elif filter == 'loc':
        locVar.set(value)
    updateText()


def updateFilters():
    #tag filter
    tagsMenu.delete(0, 'end')
    tagsMenu.add_command(label='Tags', command=lambda: setFilter('tag', 'Tags'))
    tagsMenu.add_separator()
    for i in tags:
        tagsMenu.add_command(label=i, command=lambda value=i: setFilter('tag', value))

    #location filter
    locMenu.delete(0, 'end')
    locMenu.add_command(label='Locations', command=lambda: setFilter('loc', 'Locations'))
    locMenu.add_separator()
    for i in locations:
        locMenu.add_command(label=i, command=lambda value=i: setFilter('loc', value))


def displayHistLine(line):
    color_list = ['SteelBlue1', 'turquoise1', 'paleturquoise', 'bisque2']

    #Used by inserted and deleted history only
    def highlight(tagname, end, color):
        histText['state'] = 'normal'
        histText.insert(1.0, line + '\n')
        histText.tag_add(tagname, 1.0, end)
        histText.tag_config(tagname, background=color)
        histText['state'] = 'disabled'

    if line.split()[0] == 'Inserted:': highlight('Inserted', 1.8, 'spring green')
    elif line.split()[0] == 'Deleted': highlight('Deleted', 1.7, 'firebrick2')
    elif line.split()[0] == 'Edited': #Colors items with changes
        
        left_ind = line.index(':') + 2
        right_ind = left_ind
        line_list = re.split('\|| => ',line[left_ind:-1])
        for i in range(4):
            right_ind += len(line_list[i])
        right_ind += 7
        
        histText['state'] = 'normal'
        histText.insert(1.0, line + '\n')
        histText.tag_add('Edited', 1.0, 1.6)
        histText.tag_config('Edited', background='yellow')
        
        for i in range(4):
            left_val = line_list[i]
            right_val = line_list[i+4]
            left_end_ind = left_ind + len(left_val)
            right_end_ind = right_ind + len(right_val)
            if left_val != right_val:
                histText.tag_add(f'tag{i}', f'1.{left_ind}', f'1.{left_end_ind}')
                histText.tag_config(f'tag{i}', background=color_list[i])
                histText.tag_add(f'tag{i}', f'1.{right_ind}', f'1.{right_end_ind}')
                histText.tag_config(f'tag{i}', background=color_list[i])
            left_ind = left_end_ind + 1
            right_ind = right_end_ind + 1                                                

        histText['state'] = 'disabled'

def openHistoryWindow(event):
    global histFirstOpen
    if histFirstOpen:
        for line in open(sspDir + 'text_files/change_tracker.txt'): displayHistLine(line)
        histFirstOpen = False
    historyWindow.deiconify()

#================================================= Create Main Window ===================================================
root = Tk()
root.title('Item finder')
root.resizable(0,0) #remove maximize button

ws = root.winfo_screenwidth() #get device screeen width
hs = root.winfo_screenheight() #get device screen height

rootWidth = 478
rootHeight = hs - 63

root.geometry(f'{rootWidth}x{rootHeight}+0+0') #initialize window position

#========================================== Create Withdrawn EditWindow =================================================

editWindow = Toplevel(root)
editWindow.resizable(0,0)

editWindow.geometry(f'400x300+{ws//2-200}+{hs//2-200}') #initialize window position

#create labelframes in editWindow
itemLF = LabelFrame(editWindow, text='Item')
itemLF.grid(row=0)
priceLF = LabelFrame(editWindow, text='Price')
priceLF.grid(row=1)
tagLF = LabelFrame(editWindow, text='Tag')
tagLF.grid(row=2)
locationLF = LabelFrame(editWindow, text='Location')
locationLF.grid(row=3, pady=1)

#create submitButton in editWindow
submitButton = Button(editWindow, text='Submit', bg='dim gray', activebackground='dim gray',\
    command=lambda: submitItem())
submitButton.grid(row=4)

#center widgets in editWindow
editWindow.columnconfigure(0, weight=1)
editWindow.rowconfigure(0, weight=1)
editWindow.rowconfigure(1, weight=1)
editWindow.rowconfigure(2, weight=1)
editWindow.rowconfigure(3, weight=1)
editWindow.rowconfigure(4, weight=1)

#create entries in editWindow
itemEntry = Entry(itemLF, bd=3, bg='white', fg='black', width=37, justify=CENTER)
itemEntry.grid(row=0, padx=5, pady=5)
priceEntry = Entry(priceLF, bd=3, bg='white', fg='black', width=37, justify=CENTER)
priceEntry.grid(row=0, padx=5, pady=5)

#create comboboxes in editWindow
style = ttk.Style()
style.theme_create('red', parent='alt', 
    settings = {'TCombobox': 
                {'configure': 
                    {'fieldbackground': 'red'
                    }}}
)
style.theme_create('white', parent='alt', 
    settings = {'TCombobox': 
                {'configure': 
                    {'fieldbackground': 'white'
                    }}}
)

tagCb = ttk.Combobox(tagLF, foreground='black', width=37, justify=CENTER)
tagCb.grid(row=0, padx=5, pady=5)
locationCb = ttk.Combobox(locationLF, width=37, justify=CENTER)
locationCb.grid(row=0, padx=5, pady=5)

editWindow.withdraw()

edit_add = StringVar(root)

#========================================== Create Withdrawn History Window =============================================

historyWindow = Toplevel(root, bg='gray')
historyWindow.title('History')
historyWindow.resizable(0,0) #remove maximize button

hwWidth = ws-(rootWidth+15)
hwHeight = hs-63

historyWindow.geometry(f'{hwWidth}x{hwHeight}+{ws-10}+0') #initialize window position

# historyWindow.withdraw()

histText = Text(historyWindow, bg='white', fg='black', relief=RIDGE, state=DISABLED)
histText.grid(row=0, column=0, padx=2, pady=2, sticky=N+E+S+W)
historyWindow.columnconfigure(0, weight=1)
historyWindow.rowconfigure(0, weight=1)

historyWindow.withdraw()


#============================================== Create Search & Filters =================================================

search_frame = Frame(root)
search_frame.grid(row=0)

ent = Entry(search_frame, bd=3, bg='white', fg='black', width=23)
ent.grid(row=0, column=0, padx=2)
ent.focus_set()

tags =  fetchTags()

tagsVar = StringVar(root)
tagsVar.set('Tags')
tagsOM = OptionMenu(search_frame, tagsVar, ())
tagsOM.configure(width=7)
tagsOM.grid(row=0, column=1)
tagsMenu = tagsOM['menu']

locations = fetchLocs()

locVar = StringVar(root)
locVar.set('Locations')
locationsOM = OptionMenu(search_frame, locVar, ())
locationsOM.configure(width=7)
locationsOM.grid(row=0, column=2)
locMenu = locationsOM['menu']

updateFilters()

#=================================================== Create Add Button ==================================================

button_font = tkFont.Font(size=9, weight=tkFont.BOLD)

addBut = Button(search_frame, text='+', bd=2, font=button_font, command=lambda: showEditWindow('add'))
addBut.grid(row=0, column=3, pady=1)

#================================================= Create Price List Text ===============================================

text_frame = Frame(root)
text_frame.grid(row=1)

price_list = Text(text_frame, width=59, height=57, cursor='arrow')
price_list.grid(row=0, pady=10)

price_list.insert(1.0, col_title)
price_list['state'] = 'disabled'

#=================================================  == Binds & Protocols ================================================

editWindow.protocol('WM_DELETE_WINDOW', editWindowWithdraw)
historyWindow.protocol('WM_DELETE_WINDOW', lambda: historyWindow.withdraw())

root.bind('<KeyRelease>', mapKey)
price_list.bind('<1>', selectItem)
price_list.bind('<Delete>', delItem)
root.bind('<Control-o>', openHistoryWindow)


root.mainloop()